from pathlib import Path
import tempfile, shutil
from openpyxl import load_workbook
from app.engine.fc_io_generator import generate_project_rows, row_severity
from app.engine.io_summary_engine import summarize_project
from app.export.fc_io_excel_exporter import (
    export_fc_io_workbook, build_project_export_info, default_export_filename,
)
from app.model.device import Device
from app.model.signal import Signal

def pump(tag, addrs):
    names=[('Start Command','DO',True),('Stop Command','DO',False),('Local/Remote Mode','DI',False),('Run Feedback','DI',True),('Fault Feedback','DI',True)]
    sigs=[Signal(n,io,req,True,address=addrs.get(n,'')) for n,io,req in names]
    return Device(id=tag,tag=tag,area='PRE',category='Equipment',type='Pump',description='Pump',quantity=1,signals=sigs)

# filename
assert default_export_filename(build_project_export_info())=='FC_IO_Export.xlsx'
assert default_export_filename(build_project_export_info(project_name='Demo', revision='A')).startswith('FC_IO_Demo_A')

# Test A
devices=[]
for tag, base_i, base_q in [('P-101',0,0),('P-102',1,1),('P-301',2,2),('P-302',3,3)]:
    devices.append(pump(tag,{
        'Start Command':f'Q{base_q}.0','Stop Command':f'Q{base_q}.1',
        'Local/Remote Mode':f'I{base_i}.0','Run Feedback':f'I{base_i}.1','Fault Feedback':f'I{base_i}.2',
    }))
result=generate_project_rows(devices)
assert result.total_count==20 and result.di_count==12 and result.do_count==8
ps=summarize_project(devices)
assert ps['DI']==12 and ps['DO']==8 and ps['TOTAL']==20

outdir=Path(r'D:\EEDP\.tmp_export_test')
if outdir.exists(): shutil.rmtree(outdir)
outdir.mkdir()
path=outdir/'test_a.xlsx'
export_fc_io_workbook(path, devices=devices, result=result, project_info=build_project_export_info(project_name='Test', revision='1'))
wb=load_workbook(path)
assert wb.sheetnames==['Project','Device List','FC_IO','I/O Summary','PLC Card Summary','Validation']
assert wb['Device List'].max_row==5  # header+4
assert wb['FC_IO'].max_row==21  # header+20
# filters and freeze
assert wb['FC_IO'].freeze_panes=='A2'
assert wb['FC_IO'].auto_filter.ref is not None
# Required Yes/No
assert wb['FC_IO'].cell(2,9).value in ('Yes','No')
print('A OK', result.total_count)

# Test B blank address
devices[0].signals[0].address=''
result=generate_project_rows(devices)
path=outdir/'test_b.xlsx'
export_fc_io_workbook(path, devices=devices, result=result)
wb=load_workbook(path)
# find Start Command of P-101
found=False
for r in range(2, wb['FC_IO'].max_row+1):
    if wb['FC_IO'].cell(r,4).value=='P-101' and wb['FC_IO'].cell(r,6).value=='Start Command':
        assert wb['FC_IO'].cell(r,8).value in (None,'')
        assert wb['FC_IO'].cell(r,17).value=='WARNING'
        # yellow fill
        fill=wb['FC_IO'].cell(r,1).fill.fgColor.rgb if wb['FC_IO'].cell(r,1).fill.fgColor else ''
        assert 'FFF9C4' in str(fill).upper() or 'FFFF9C4' in str(fill).upper()
        found=True
        break
assert found
assert any(wb['Validation'].cell(r,2).value=='MISSING_PLC_ADDRESS' for r in range(2, wb['Validation'].max_row+1))
print('B OK')

# Test C duplicate
devices[0].signals[0].address='Q0.0'
devices[0].signals[3].address='I0.0'  # Run same as Local
# Local is already I0.0
result=generate_project_rows([devices[0]])
assert result.error_count>=2
path=outdir/'test_c.xlsx'
export_fc_io_workbook(path, devices=[devices[0]], result=result)
wb=load_workbook(path)
reds=0
for r in range(2, wb['FC_IO'].max_row+1):
    if wb['FC_IO'].cell(r,8).value and str(wb['FC_IO'].cell(r,8).value).upper()=='I0.0':
        assert wb['FC_IO'].cell(r,17).value=='ERROR'
        fill=str(wb['FC_IO'].cell(r,1).fill.fgColor.rgb).upper()
        assert 'FFCDD2' in fill
        reds+=1
assert reds==2
assert any(wb['Validation'].cell(r,2).value=='DUPLICATE_PLC_ADDRESS' for r in range(2, wb['Validation'].max_row+1))
print('C OK')

# Test D: export uses full result not filtered - already exporting result.rows full
assert len(result.rows)==generate_project_rows([devices[0]]).total_count

# UI button enabled
import os
os.environ.setdefault('QT_QPA_PLATFORM','offscreen')
from PySide6.QtWidgets import QApplication
from app.ui.dialogs.fc_io_preview_dialog import FCIOPreviewDialog
app=QApplication([])
dlg=FCIOPreviewDialog(generate_project_rows(devices))
assert dlg.export_button.isEnabled()
assert hasattr(dlg, 'export_excel_requested')
print('D/UI OK')
print('ALL PASS')
print('workbook', path)