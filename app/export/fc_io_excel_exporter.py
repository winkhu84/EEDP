"""Export FC_IO generation results to a formatted Excel workbook."""

from __future__ import annotations

from collections.abc import Sequence
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.common.constants import APP_VERSION
from app.engine.fc_io_generator import (
    FCIOGenerationResult,
    FCIOIssue,
    row_severity,
    sort_devices,
)
from app.engine.io_summary_engine import summarize_device, summarize_project
from app.engine.plc_card_calculator import (
    PlcCardRequirement,
    calculate_project_cards,
)
from app.model.device import Device
from app.model.plc_card_config import PlcCardConfig, default_plc_card_configurations

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_DATA_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
_CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
_THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
_WARNING_FILL = PatternFill("solid", fgColor="FFF9C4")
_ERROR_FILL = PatternFill("solid", fgColor="FFCDD2")
_ALT_FILL = PatternFill("solid", fgColor="F5F5F5")
_SECTION_FONT = Font(bold=True, size=12)

_MAX_WIDTHS = {
    "Device Description": 35,
    "Signal Description": 40,
    "Remark": 40,
    "Message": 60,
    "Validation Message": 60,
}


@dataclass(frozen=True)
class ProjectExportInfo:
    """Project metadata written to the Project worksheet."""

    customer: str = ""
    project_name: str = ""
    revision: str = ""
    plc: str = ""
    cpu: str = ""
    designer: str = ""
    date: str = ""
    export_date: str = ""
    eos_version: str = ""
    ecs_version: str = ""
    library_version: str = ""


def build_project_export_info(
    *,
    customer: str = "",
    project_name: str = "",
    revision: str = "",
    plc: str = "",
    cpu: str = "",
    designer: str = "",
    project_date: str = "",
    eos_version: str | None = None,
    ecs_version: str = "",
    library_version: str = "",
) -> ProjectExportInfo:
    """Build export metadata; blanks are allowed."""
    return ProjectExportInfo(
        customer=customer.strip(),
        project_name=project_name.strip(),
        revision=revision.strip(),
        plc=plc.strip(),
        cpu=cpu.strip(),
        designer=designer.strip(),
        date=project_date.strip(),
        export_date=date.today().isoformat(),
        eos_version=(APP_VERSION if eos_version is None else eos_version),
        ecs_version=ecs_version.strip(),
        library_version=library_version.strip(),
    )


def default_export_filename(info: ProjectExportInfo) -> str:
    """Return FC_IO_<ProjectName>_<Revision>.xlsx or FC_IO_Export.xlsx."""
    name = info.project_name.strip()
    revision = info.revision.strip()
    if name and revision:
        return f"FC_IO_{_safe_filename(name)}_{_safe_filename(revision)}.xlsx"
    if name:
        return f"FC_IO_{_safe_filename(name)}.xlsx"
    return "FC_IO_Export.xlsx"


def _safe_filename(text: str) -> str:
    cleaned = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in text.strip())
    return cleaned.strip("_") or "Export"


def export_fc_io_workbook(
    file_path: str | Path,
    *,
    devices: Sequence[Device],
    result: FCIOGenerationResult,
    project_info: ProjectExportInfo | None = None,
    card_configurations: Sequence[PlcCardConfig] | None = None,
) -> Path:
    """Write a six-sheet FC_IO workbook. Raises OSError on file failures."""
    path = Path(file_path)
    info = project_info or build_project_export_info()
    configs = (
        tuple(card_configurations)
        if card_configurations is not None
        else default_plc_card_configurations()
    )
    project_summary = summarize_project(devices)
    plc_cards = calculate_project_cards(project_summary, configs)

    workbook = Workbook()
    # Remove default sheet after creating named ones.
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    _write_project_sheet(workbook.create_sheet("Project"), info)
    _write_device_list_sheet(workbook.create_sheet("Device List"), devices)
    _write_fc_io_sheet(workbook.create_sheet("FC_IO"), result)
    _write_io_summary_sheet(
        workbook.create_sheet("IO Summary"),
        devices,
        project_summary,
    )
    _write_plc_card_sheet(workbook.create_sheet("PLC Card Summary"), plc_cards)
    _write_validation_sheet(workbook.create_sheet("Validation"), result)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def _style_header(worksheet: Worksheet, headers: Sequence[str]) -> None:
    worksheet.append(list(headers))
    for column, _header in enumerate(headers, start=1):
        cell = worksheet.cell(1, column)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = (
        f"A1:{get_column_letter(len(headers))}1"
    )


def _finalize_table(
    worksheet: Worksheet,
    headers: Sequence[str],
    row_count: int,
) -> None:
    last_row = max(1, row_count)
    worksheet.auto_filter.ref = (
        f"A1:{get_column_letter(len(headers))}{last_row}"
    )
    for column, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in range(2, last_row + 1):
            value = worksheet.cell(row, column).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        cap = _MAX_WIDTHS.get(header, 28)
        width = min(max(max_len + 2, 10), cap)
        worksheet.column_dimensions[get_column_letter(column)].width = width


def _apply_data_cell(
    cell,
    *,
    fill: PatternFill | None = None,
    align: Alignment | None = None,
    bold: bool = False,
) -> None:
    cell.border = _THIN
    cell.alignment = align or _DATA_ALIGN
    if fill is not None:
        cell.fill = fill
    if bold:
        cell.font = Font(bold=True)


def _write_project_sheet(worksheet: Worksheet, info: ProjectExportInfo) -> None:
    headers = ("Item", "Value")
    _style_header(worksheet, headers)
    items = (
        ("Customer", info.customer),
        ("Project Name", info.project_name),
        ("Revision", info.revision),
        ("PLC", info.plc),
        ("CPU", info.cpu),
        ("Designer", info.designer),
        ("Date", info.date),
        ("Export Date", info.export_date),
        ("EOS Version", info.eos_version),
        ("ECS Version", info.ecs_version),
        ("Library Version", info.library_version),
    )
    for index, (item, value) in enumerate(items, start=2):
        worksheet.cell(index, 1, item)
        worksheet.cell(index, 2, value)
        fill = _ALT_FILL if index % 2 == 0 else None
        for column in (1, 2):
            _apply_data_cell(worksheet.cell(index, column), fill=fill)
    _finalize_table(worksheet, headers, 1 + len(items))


def _write_device_list_sheet(
    worksheet: Worksheet,
    devices: Sequence[Device],
) -> None:
    headers = (
        "No.",
        "Area",
        "Category",
        "Device Type",
        "Device Tag",
        "Description",
        "DI Count",
        "DO Count",
        "AI Count",
        "AO Count",
        "Total Signals",
        "DI Start Address",
        "DO Start Address",
        "AI Start Address",
        "AO Start Address",
    )
    _style_header(worksheet, headers)
    sorted_devices = sort_devices(devices)
    for index, device in enumerate(sorted_devices, start=1):
        summary = summarize_device(device)
        values = (
            index,
            device.area,
            device.category,
            device.type,
            device.tag,
            device.description,
            summary["DI"],
            summary["DO"],
            summary["AI"],
            summary["AO"],
            summary["TOTAL"],
            device.di_start_address,
            device.do_start_address,
            device.ai_start_address,
            device.ao_start_address,
        )
        excel_row = index + 1
        fill = _ALT_FILL if index % 2 == 0 else None
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(excel_row, column, value)
            _apply_data_cell(cell, fill=fill)
    _finalize_table(worksheet, headers, 1 + len(sorted_devices))


def _validation_status(severity: str | None) -> str:
    if severity == "error":
        return "ERROR"
    if severity == "warning":
        return "WARNING"
    return "OK"


def _write_fc_io_sheet(worksheet: Worksheet, result: FCIOGenerationResult) -> None:
    headers = (
        "No.",
        "Area",
        "Device Type",
        "Device Tag",
        "Device Description",
        "Signal Name",
        "I/O Type",
        "PLC Address",
        "Required",
        "Signal Description",
        "Remark",
        "Rack",
        "Slot",
        "Channel",
        "Terminal",
        "Cable",
        "Validation Status",
    )
    _style_header(worksheet, headers)

    for index, row in enumerate(result.rows, start=1):
        severity = row_severity(row, result.warnings, result.errors)
        status = _validation_status(severity)
        values = (
            index,
            row.area,
            row.device_type,
            row.device_tag,
            row.device_description,
            row.signal_name,
            row.io_type,
            row.plc_address,
            "Yes" if row.required else "No",
            row.signal_description,
            row.remark,
            row.rack,
            row.slot,
            row.channel,
            row.terminal,
            row.cable,
            status,
        )
        excel_row = index + 1
        if severity == "error":
            fill = _ERROR_FILL
        elif severity == "warning":
            fill = _WARNING_FILL
        elif index % 2 == 0:
            fill = _ALT_FILL
        else:
            fill = None

        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(excel_row, column, value)
            header = headers[column - 1]
            if header == "PLC Address":
                _apply_data_cell(
                    cell,
                    fill=fill,
                    align=_CENTER_ALIGN,
                    bold=bool(str(value).strip()),
                )
            elif header in {"No.", "I/O Type", "Required", "Validation Status"}:
                _apply_data_cell(cell, fill=fill, align=_CENTER_ALIGN)
            else:
                _apply_data_cell(cell, fill=fill)

    _finalize_table(worksheet, headers, 1 + len(result.rows))


def _write_io_summary_sheet(
    worksheet: Worksheet,
    devices: Sequence[Device],
    project_summary: dict[str, int],
) -> None:
    worksheet["A1"] = "Device I/O Summary"
    worksheet["A1"].font = _SECTION_FONT

    device_headers = ("Device Tag", "DI", "DO", "AI", "AO", "Total")
    start_row = 3
    for column, header in enumerate(device_headers, start=1):
        cell = worksheet.cell(start_row, column, header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN

    sorted_devices = sort_devices(devices)
    for index, device in enumerate(sorted_devices, start=1):
        summary = summarize_device(device)
        values = (
            device.tag,
            summary["DI"],
            summary["DO"],
            summary["AI"],
            summary["AO"],
            summary["TOTAL"],
        )
        excel_row = start_row + index
        fill = _ALT_FILL if index % 2 == 0 else None
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(excel_row, column, value)
            align = _CENTER_ALIGN if column > 1 else _DATA_ALIGN
            _apply_data_cell(cell, fill=fill, align=align)

    project_title_row = start_row + len(sorted_devices) + 3
    worksheet.cell(project_title_row, 1, "Project I/O Summary").font = _SECTION_FONT

    project_header_row = project_title_row + 2
    for column, header in enumerate(("I/O Type", "Count"), start=1):
        cell = worksheet.cell(project_header_row, column, header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN

    for offset, key in enumerate(("DI", "DO", "AI", "AO", "Total"), start=1):
        excel_row = project_header_row + offset
        worksheet.cell(excel_row, 1, key if key != "Total" else "Total")
        worksheet.cell(
            excel_row,
            2,
            project_summary.get(key.upper() if key != "Total" else "TOTAL", 0),
        )
        fill = _ALT_FILL if offset % 2 == 0 else None
        for column in (1, 2):
            _apply_data_cell(
                worksheet.cell(excel_row, column),
                fill=fill,
                align=_CENTER_ALIGN if column == 2 else _DATA_ALIGN,
                bold=(key == "Total"),
            )

    worksheet.freeze_panes = "A4"
    worksheet.column_dimensions["A"].width = 18
    for column in range(2, 7):
        worksheet.column_dimensions[get_column_letter(column)].width = 10


def _write_plc_card_sheet(
    worksheet: Worksheet,
    requirements: Sequence[PlcCardRequirement],
) -> None:
    headers = (
        "I/O Type",
        "Module",
        "Used Channels",
        "Channels per Card",
        "Required Cards",
        "Total Channels",
        "Spare Channels",
    )
    _style_header(worksheet, headers)
    for index, item in enumerate(requirements, start=1):
        values = (
            item.io_type,
            item.module_name,
            item.used_channels,
            item.channels_per_card,
            item.required_cards,
            item.total_channels,
            item.spare_channels,
        )
        excel_row = index + 1
        fill = _ALT_FILL if index % 2 == 0 else None
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(excel_row, column, value)
            align = _CENTER_ALIGN if column != 2 else _DATA_ALIGN
            _apply_data_cell(cell, fill=fill, align=align)
    _finalize_table(worksheet, headers, 1 + len(requirements))


def _write_validation_sheet(
    worksheet: Worksheet,
    result: FCIOGenerationResult,
) -> None:
    headers = (
        "Severity",
        "Code",
        "Area",
        "Device Tag",
        "Signal Name",
        "I/O Type",
        "PLC Address",
        "Message",
    )
    _style_header(worksheet, headers)

    issues: list[FCIOIssue] = list(result.errors) + list(result.warnings)
    if not issues:
        values = (
            "INFO",
            "",
            "",
            "",
            "",
            "",
            "",
            "No validation issues detected.",
        )
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(2, column, value)
            _apply_data_cell(cell)
        _finalize_table(worksheet, headers, 2)
        return

    for index, issue in enumerate(issues, start=1):
        values = (
            issue.severity.upper(),
            issue.code,
            issue.area,
            issue.device_tag,
            issue.signal_name,
            issue.io_type,
            issue.plc_address,
            issue.message,
        )
        excel_row = index + 1
        if issue.severity == "error":
            fill = _ERROR_FILL
        elif issue.severity == "warning":
            fill = _WARNING_FILL
        else:
            fill = None
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(excel_row, column, value)
            _apply_data_cell(cell, fill=fill)
    _finalize_table(worksheet, headers, 1 + len(issues))
