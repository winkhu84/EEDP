from pathlib import Path
import shutil
from openpyxl import load_workbook
from app.engine.fc_io_generator import generate_project_rows
from app.engine.io_summary_engine import summarize_project
from app.export.fc_io_excel_exporter import export_fc_io_workbook, build_project_export_info, default_export_filename
from app.model.device import Device
from app.model.signal import Signal

def pump(tag, addrs):
    names=[('Start Command','DO',True),('Stop Command','DO',False),('Local/Remote Mode','DI',False),('Run Feedback','DI',True),('Fault Feedback','DI',True)]
    sigs=[Signal(n,io,req,True,address=addrs.get(n,'')) for n,io,req in names]
    return Device(id=tag,tag=tag,area='PRE',category='Equipment',type='Pump',description='Pump',quantity=1,signals=sigs)

assert default_export_filename(build_project_export_info())=='FC_IO_Export.xlsx'
devices=[]
for tag, bi, bq in [('P-101',0,0),('P-102',1,1),('P-301',2,2),('P-302',3,3)]:
    devices.append(pump(tag,{'Start Command':f'Q{bq}.0','Stop Command':f'Q{bq}.1','Local/Remote Mode':f'I{bi}.0','Run Feedback':f'I{bi}.1','Fault Feedback':f'I{bi}.2'}))
result=generate_project_rows(devices)
assert result.total_count==20 and result.di_count==12 and result.do_count==8
outdir=Path(r'D:\EEDP\.tmp_export_test');
if outdir.exists(): shutil.rmtree(outdir)
outdir.mkdir(); path=outdir/'test_a.xlsx'
export_fc_io_workbook(path, devices=devices, result=result, project_info=build_project_export_info(project_name='Test', revision='1'))
wb=load_workbook(path)
assert wb.sheetnames==['Project','Device List','FC_IO','IO Summary','PLC Card Summary','Validation'], wb.sheetnames
assert wb['Device List'].max_row==5 and wb['FC_IO'].max_row==21
assert wb['FC_IO'].freeze_panes=='A2'
print('A OK')
devices[0].signals[0].address=''
result=generate_project_rows(devices)
path=outdir/'test_b.xlsx'; export_fc_io_workbook(path, devices=devices, result=result)
wb=load_workbook(path)
found=False
for r in range(2, wb['FC_IO'].max_row+1):
    if wb['FC_IO'].cell(r,4).value=='P-101' and wb['FC_IO'].cell(r,6).value=='Start Command':
        assert wb['FC_IO'].cell(r,8).value in (None,'')
        assert wb['FC_IO'].cell(r,17).value=='WARNING'
        fill=str(wb['FC_IO'].cell(r,1).fill.fgColor.rgb).upper()
        assert 'FFF9C4' in fill
        found=True
assert found and any(wb['Validation'].cell(r,2).value=='MISSING_PLC_ADDRESS' for r in range(2, wb['Validation'].max_row+1))
print('B OK')
devices[0].signals[0].address='Q0.0'
devices[0].signals[3].address='I0.0'
result=generate_project_rows([devices[0]])
path=outdir/'test_c.xlsx'; export_fc_io_workbook(path, devices=[devices[0]], result=result)
wb=load_workbook(path)
reds=0
for r in range(2, wb['FC_IO'].max_row+1):
    if str(wb['FC_IO'].cell(r,8).value or '').upper()=='I0.0':
        assert wb['FC_IO'].cell(r,17).value=='ERROR'
        assert 'FFCDD2' in str(wb['FC_IO'].cell(r,1).fill.fgColor.rgb).upper()
        reds+=1
assert reds==2
print('C OK')
import os
os.environ.setdefault('QT_QPA_PLATFORM','offscreen')
from PySide6.QtWidgets import QApplication
from app.ui.dialogs.fc_io_preview_dialog import FCIOPreviewDialog
app=QApplication([])
dlg=FCIOPreviewDialog(generate_project_rows(devices))
assert dlg.export_button.isEnabled()
print('ALL PASS')
